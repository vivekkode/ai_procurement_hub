"""
generic_parser.py
-----------------
A single parser that can handle ANY supplier file in ANY format
by reading a supplier config file instead of having format-specific
logic hardcoded.

This replaces the 4 individual parsers (parse_valero, parse_exxon,
parse_marathon, parse_pemex) for known suppliers and handles new
suppliers automatically when a config is provided.

How it works:
    1. Receives a file path + a supplier config dict
    2. Uses the config to know: what format, where is the date,
       which columns are price/product/terminal, what unit is used
    3. Extracts data and returns a standard schema DataFrame —
       identical output to the original parsers

Adding a new supplier:
    - Write a config file in config/suppliers/
    - Drop supplier files in data/raw/<supplier_name>/
    - Run the pipeline — zero code changes needed

Output schema (identical to all_suppliers.csv):
    date            | YYYY-MM-DD
    supplier        | from config
    terminal_id     | extracted per config rules
    terminal_name   | extracted per config rules
    state           | extracted where available
    country         | from config
    product_type    | normalized via config product_map
    price_mxn_per_l | converted to MXN/L if needed
    contract_type   | from config
    source_file     | original filename
    price_flag      | True if outside guardrail range

Usage:
    from src.ingestion.generic_parser import GenericParser
    from config.suppliers.exxon_config import CONFIG

    parser = GenericParser(CONFIG)
    df = parser.parse_file("data/raw/exxon/exxon_20240101.xlsx")
    df = parser.parse_folder("data/raw/exxon/")
"""

import re
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd

from src.ingestion.format_detector import detect_format, FileFormat

logger = logging.getLogger(__name__)

# Standard output columns — never changes regardless of supplier
OUTPUT_COLUMNS = [
    "date", "supplier", "terminal_id", "terminal_name",
    "state", "country", "product_type", "price_mxn_per_l",
    "contract_type", "source_file", "price_flag",
]


# ---------------------------------------------------------------------------
# GenericParser
# ---------------------------------------------------------------------------

class GenericParser:
    """
    Parses supplier pricing files using a config dict.

    One instance per supplier. The config completely describes how to
    extract data from that supplier's files — no hardcoded logic here.
    """

    def __init__(self, config: dict):
        self.config = config
        self.supplier_name = config["supplier_name"]
        self.logger = logging.getLogger(
            f"{__name__}.{self.supplier_name}"
        )

    # ── Public API ──────────────────────────────────────────────────────

    def parse_file(self, filepath: str) -> pd.DataFrame:
        """
        Parse a single supplier file into a standard schema DataFrame.

        Args:
            filepath: Path to any supplier file

        Returns:
            pd.DataFrame with OUTPUT_COLUMNS, empty on failure
        """
        filepath = Path(filepath)

        if not filepath.exists():
            self.logger.error("File not found: %s", filepath)
            return pd.DataFrame()

        # Detect format — trust magic bytes over extension
        fmt = detect_format(str(filepath))

        self.logger.debug("Parsing %s as %s", filepath.name, fmt)

        # Route to the correct extraction method
        try:
            if fmt == FileFormat.XLSX or fmt == FileFormat.XLS:
                records = self._extract_xlsx(filepath)
            elif fmt == FileFormat.HTML:
                records = self._extract_html(filepath)
            elif fmt == FileFormat.TXT:
                records = self._extract_txt(filepath)
            elif fmt == FileFormat.PDF:
                records = self._extract_pdf(filepath)
            elif fmt == FileFormat.CSV:
                records = self._extract_csv(filepath)
            else:
                self.logger.warning(
                    "Format %s not directly handled for %s — "
                    "route to LLM fallback",
                    fmt, filepath.name
                )
                return pd.DataFrame()
        except Exception as e:
            self.logger.error(
                "Extraction failed for %s: %s", filepath.name, e
            )
            return pd.DataFrame()

        if not records:
            self.logger.warning("No records extracted from: %s", filepath.name)
            return pd.DataFrame()

        df = self._to_dataframe(records, filepath.name)
        self.logger.info(
            "Parsed %s → %d records | %d terminals",
            filepath.name, len(df),
            df["terminal_id"].nunique() if not df.empty else 0
        )
        return df

    def parse_folder(self, folder_path: str) -> pd.DataFrame:
        """
        Parse all compatible files in a folder.

        Args:
            folder_path: Path to supplier data folder

        Returns:
            Combined DataFrame sorted by date, empty if nothing found
        """
        folder = Path(folder_path)
        if not folder.exists():
            self.logger.error("Folder not found: %s", folder_path)
            return pd.DataFrame()

        extensions = self.config.get("file_extensions", [])
        all_files = []
        for ext in extensions:
            all_files.extend(sorted(folder.glob(f"*{ext}")))

        # Also scan subdirectories one level deep (e.g. pemex/reportes_pdf/)
        for subdir in folder.iterdir():
            if subdir.is_dir():
                for ext in extensions:
                    all_files.extend(sorted(subdir.glob(f"*{ext}")))

        if not all_files:
            self.logger.warning(
                "No files with extensions %s found in: %s",
                extensions, folder_path
            )
            return pd.DataFrame()

        self.logger.info(
            "Found %d files to parse in: %s", len(all_files), folder_path
        )

        frames = []
        failed = 0
        for f in all_files:
            df = self.parse_file(str(f))
            if df.empty:
                failed += 1
            else:
                frames.append(df)

        if not frames:
            self.logger.error("All files failed in: %s", folder_path)
            return pd.DataFrame()

        combined = pd.concat(frames, ignore_index=True)
        combined = combined.drop_duplicates(
            subset=["date", "terminal_id", "product_type"]
        )
        combined = combined.sort_values("date").reset_index(drop=True)

        self.logger.info(
            "Total: %d records | %d terminals | %d files failed",
            len(combined),
            combined["terminal_id"].nunique(),
            failed,
        )
        return combined

    # ── Format-specific extractors ──────────────────────────────────────

    def _extract_xlsx(self, filepath: Path) -> list:
        """Extract records from Excel files using config column_map."""
        import openpyxl

        cfg = self.config
        sheet_name = cfg.get("sheet_name")
        sheet_index = cfg.get("sheet_index", 0)
        header_row = cfg.get("header_row", 0)
        data_start = cfg.get("data_start", header_row + 1)

        try:
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
        except Exception as e:
            self.logger.error("Cannot open Excel file %s: %s", filepath.name, e)
            return []

        # Get sheet by name, fall back to index
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            if sheet_name:
                self.logger.warning(
                    "Sheet '%s' not found in %s — using sheet index %d",
                    sheet_name, filepath.name, sheet_index
                )
            ws = wb.worksheets[sheet_index]

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row:
            return []

        # Extract date
        date_str = self._extract_date_xlsx(rows, filepath)
        if not date_str:
            self.logger.warning("No date found in: %s", filepath.name)
            return []

        # Build column map from header row
        headers = [
            str(c).strip().lower() if c else ""
            for c in rows[header_row]
        ]
        col_map = self._build_col_map(headers)

        # Extract records from data rows
        records = []
        for row in rows[data_start:]:
            if not any(row):
                continue

            record = self._extract_row_xlsx(row, col_map, date_str, filepath.name)
            if record:
                records.append(record)

        return records

    def _extract_date_xlsx(self, rows: list, filepath: Path) -> str:
        """Extract date from Excel file — title row or filename."""
        cfg = self.config
        date_source = cfg.get("date_source", "filename")
        patterns = cfg.get("date_patterns", [r"(\d{8})"])

        if date_source == "title_row" and rows:
            # Look in first row for date string
            title = " ".join(str(c) for c in rows[0] if c)
            for pattern in patterns:
                match = re.search(pattern, title)
                if match:
                    fmt = cfg.get("date_format_primary", "%d/%m/%Y")
                    try:
                        return datetime.strptime(match.group(1), fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass

        # Fallback: filename
        return self._extract_date_filename(filepath.name)

    def _extract_row_xlsx(self, row: tuple, col_map: dict,
                          date_str: str, source_file: str) -> dict:
        """Extract one record from an Excel row using col_map."""
        cfg = self.config

        def get(field, default=""):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else default
            return default

        terminal_raw = get("terminal")
        if not terminal_raw or terminal_raw.lower() in ("none", ""):
            return {}

        terminal_info = self._parse_terminal(terminal_raw)
        product_raw   = get("product")
        product_type  = self._normalize_product(product_raw)

        price_raw = get("price")
        price     = self._parse_price(price_raw)
        if price is None:
            return {}

        return {
            "date":            date_str,
            "supplier":        cfg["supplier_name"],
            "terminal_id":     terminal_info.get("terminal_id", terminal_raw),
            "terminal_name":   terminal_info.get("terminal_name", terminal_raw),
            "state":           terminal_info.get("state", ""),
            "country":         terminal_info.get("country", cfg.get("country", "MX")),
            "product_raw":     product_raw,
            "product_type":    product_type,
            "price_mxn_per_l": self._convert_price(price),
            "contract_type":   cfg.get("contract_type", ""),
            "source_file":     source_file,
        }

    def _extract_html(self, filepath: Path) -> list:
        """Extract records from HTML files using config selectors."""
        from bs4 import BeautifulSoup

        cfg = self.config
        selectors = cfg.get("selectors", {})

        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                soup = BeautifulSoup(f, "html.parser")
        except Exception as e:
            self.logger.error("Cannot read HTML %s: %s", filepath.name, e)
            return []

        date_str = self._extract_date_filename(filepath.name)
        if not date_str:
            self.logger.warning("No date found in: %s", filepath.name)
            return []

        card_sel = selectors.get("card", {})
        cards = soup.find_all(
            card_sel.get("tag", "div"),
            class_=card_sel.get("class")
        )

        records = []
        for card in cards:
            header_sel = selectors.get("header", {})
            header_div = card.find(
                header_sel.get("tag", "div"),
                class_=header_sel.get("class")
            )
            if not header_div:
                continue

            spans = header_div.find_all("span")
            header_text = spans[0].get_text(strip=True) if spans else \
                          header_div.get_text(strip=True)
            header_text = header_text.replace("Save / Print", "").strip()

            terminal_info = self._parse_terminal(header_text)

            table = card.find("table")
            if not table:
                continue

            tbody = table.find("tbody")
            data_rows = tbody.find_all("tr") if tbody else table.find_all("tr")[1:]

            for tr in data_rows:
                cells = tr.find_all("td")
                if len(cells) < 3:
                    continue

                contract_type = cells[0].get_text(strip=True) or \
                                cfg.get("contract_type", "")
                product_raw   = cells[1].get_text(strip=True)
                product_type  = self._normalize_product(product_raw)

                price_idx = cfg.get("price_column_index", 2)
                price_raw = cells[price_idx].get_text(strip=True) \
                            if price_idx < len(cells) else ""
                price = self._parse_price(price_raw)
                if price is None:
                    continue

                records.append({
                    "date":            date_str,
                    "supplier":        cfg["supplier_name"],
                    "terminal_id":     terminal_info.get("terminal_id", ""),
                    "terminal_name":   terminal_info.get("terminal_name", header_text),
                    "state":           terminal_info.get("state", ""),
                    "country":         terminal_info.get("country", cfg.get("country", "MX")),
                    "product_raw":     product_raw,
                    "product_type":    product_type,
                    "price_mxn_per_l": self._convert_price(price),
                    "contract_type":   contract_type,
                    "source_file":     filepath.name,
                })

        return records

    def _extract_txt(self, filepath: Path) -> list:
        """Extract records from plain-text email files using config labels."""
        cfg = self.config

        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        except Exception as e:
            self.logger.error("Cannot read TXT %s: %s", filepath.name, e)
            return []

        date_str = self._extract_date_filename(filepath.name)
        terminal_info = self._parse_terminal_from_filename(filepath.name)

        if not date_str:
            self.logger.warning("No date in: %s", filepath.name)
            return []

        # Find the price row using configured labels
        price_labels = cfg.get("price_labels", ["Invoice Price"])
        product_order = cfg.get("product_order", ["Regular", "Premium", "Diesel"])
        prices = []

        for line in content.splitlines():
            line_stripped = line.strip()
            for label in price_labels:
                if line_stripped.startswith(label):
                    prices = self._extract_numbers(line_stripped)
                    break
            if prices:
                break

        if not prices:
            self.logger.warning("No price row found in: %s", filepath.name)
            return []

        records = []
        for i, product_type in enumerate(product_order):
            if i >= len(prices):
                break
            price = prices[i]
            records.append({
                "date":            date_str,
                "supplier":        cfg["supplier_name"],
                "terminal_id":     terminal_info.get("terminal_id", ""),
                "terminal_name":   terminal_info.get("terminal_name", ""),
                "state":           terminal_info.get("state", ""),
                "country":         terminal_info.get("country", cfg.get("country", "MX")),
                "product_raw":     f"UNBRANDED {product_type.upper()}",
                "product_type":    product_type,
                "price_mxn_per_l": self._convert_price(price),
                "contract_type":   cfg.get("contract_type", ""),
                "source_file":     filepath.name,
            })

        return records

    def _extract_pdf(self, filepath: Path) -> list:
        """Extract records from PDF files using config text patterns."""
        import pdfminer.high_level

        cfg = self.config

        try:
            text = pdfminer.high_level.extract_text(str(filepath))
        except Exception as e:
            self.logger.error("Cannot read PDF %s: %s", filepath.name, e)
            return []

        if not text:
            self.logger.warning("Empty PDF: %s", filepath.name)
            return []

        # Extract date range from content
        date_start, date_end = self._extract_pdf_dates(text, filepath.name)

        pages = text.split("\f")
        all_records = []

        for page in pages:
            page = page.strip()
            if not page:
                continue
            page_records = self._extract_pdf_page(
                page, date_start, date_end, filepath.name
            )
            all_records.extend(page_records)

        return all_records

    def _extract_pdf_page(self, page_text: str, date_start: str,
                          date_end: str, source_file: str) -> list:
        """Parse one PDF page into price records using config patterns."""
        cfg = self.config
        skip_words = set(cfg.get("skip_words", []))
        skip_subs  = cfg.get("skip_substrings", [])
        price_pat  = cfg.get("price_pattern", r"^\d{2,3},\d{3}\.\d{4}$")

        product_type = self._detect_pdf_product(page_text)
        if not product_type:
            return []

        lines    = [l.strip() for l in page_text.splitlines() if l.strip()]
        regions  = []
        prices   = []

        for line in lines:
            if re.match(price_pat, line):
                prices.append(float(line.replace(",", "")))
            elif (
                line.upper() == line
                and not any(c.isdigit() for c in line)
                and len(line) > 2
                and line not in skip_words
                and not any(s in line for s in skip_subs)
            ):
                regions.append(line)

        count = min(len(regions), len(prices))
        if len(regions) != len(prices):
            self.logger.debug(
                "Region/price mismatch: %d regions, %d prices — page product: %s",
                len(regions), len(prices), product_type
            )

        records = []
        factor = cfg.get("conversion_factor", 1.0)

        for i in range(count):
            price_raw = prices[i]
            price_l   = round(price_raw / factor, 6)

            records.append({
                "date":            date_start,
                "supplier":        cfg["supplier_name"],
                "terminal_id":     regions[i],
                "terminal_name":   regions[i].title(),
                "state":           "",
                "country":         cfg.get("country", "MX"),
                "product_raw":     regions[i],
                "product_type":    product_type,
                "price_mxn_per_l": price_l,
                "contract_type":   cfg.get("contract_type", ""),
                "source_file":     source_file,
            })

        return records

    def _extract_csv(self, filepath: Path) -> list:
        """Extract records from CSV files using config column_map."""
        cfg = self.config
        try:
            df = pd.read_csv(filepath, encoding="utf-8-sig")
        except Exception as e:
            self.logger.error("Cannot read CSV %s: %s", filepath.name, e)
            return []

        headers = [str(c).strip().lower() for c in df.columns]
        col_map = self._build_col_map(headers)
        date_str = self._extract_date_filename(filepath.name)

        records = []
        for _, row in df.iterrows():
            record = self._extract_row_xlsx(
                tuple(row), col_map, date_str, filepath.name
            )
            if record:
                records.append(record)
        return records

    # ── Helper methods ──────────────────────────────────────────────────

    def _build_col_map(self, headers: list) -> dict:
        """
        Map standard field names to column indices using config column_map.

        For each standard field (terminal, product, price, etc.) the config
        provides a list of possible column name keywords in any language.
        We search each header for each keyword (case-insensitive substring)
        and return the index of the first match.

        This is what makes the parser language-agnostic — Spanish column
        names like "Precio Facturación" and English "Invoice Price" both
        map to the same "price" field if both keywords are in the config.
        """
        column_map_config = self.config.get("column_map", {})
        result = {}

        for field, keywords in column_map_config.items():
            for i, header in enumerate(headers):
                for keyword in keywords:
                    if keyword.lower() in header.lower():
                        result[field] = i
                        break
                if field in result:
                    break

        return result

    def _parse_terminal(self, raw: str) -> dict:
        """Parse terminal string using config patterns."""
        cfg = self.config
        if not raw:
            return {"terminal_id": "", "terminal_name": "", "state": "", "country": cfg.get("country", "MX")}

        pattern = cfg.get("terminal_pattern")
        groups  = cfg.get("terminal_groups", {})

        if pattern:
            match = re.match(pattern, raw.strip(), re.IGNORECASE)
            if match:
                result = {
                    "terminal_id":   match.group(groups["terminal_id"]).strip()   if "terminal_id"   in groups else "",
                    "terminal_name": match.group(groups["terminal_name"]).strip() if "terminal_name" in groups else raw,
                    "state":         match.group(groups["state"]).strip()         if "state"         in groups else "",
                    "country":       match.group(groups["country"]).strip()       if "country"       in groups else cfg.get("country", "MX"),
                }
                return result

        # Try fallback pattern
        fallback = cfg.get("terminal_pattern_fallback")
        fb_groups = cfg.get("terminal_groups_fallback", {})
        if fallback:
            match = re.match(fallback, raw.strip())
            if match:
                name = match.group(fb_groups.get("terminal_name", 1)).strip()
                tid  = match.group(fb_groups.get("terminal_id", 2)).strip()
                country = "US" if any(
                    us in name.upper()
                    for us in ["EL PASO", "HARLINGEN", "BROWNSVILLE"]
                ) else cfg.get("country", "MX")
                return {
                    "terminal_id":   tid,
                    "terminal_name": name,
                    "state":         "",
                    "country":       country,
                }

        return {"terminal_id": "", "terminal_name": raw, "state": "", "country": cfg.get("country", "MX")}

    def _parse_terminal_from_filename(self, filename: str) -> dict:
        """Extract terminal info from filename (used by Marathon)."""
        cfg = self.config
        if not cfg.get("terminal_from_filename"):
            return {}

        pattern = cfg.get("terminal_pattern", "")
        groups  = cfg.get("terminal_groups", {})

        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            country_code = match.group(groups.get("country_code", 3))
            return {
                "terminal_id":   match.group(groups.get("terminal_id", 2)),
                "terminal_name": match.group(groups.get("terminal_name", 4)).replace("_", " "),
                "state":         "",
                "country":       country_code,
            }
        return {}

    def _normalize_product(self, raw: str) -> str:
        """Normalize product string to Regular/Premium/Diesel using config map."""
        if not raw:
            return "Unknown"
        lower = str(raw).lower()
        product_map = self.config.get("product_map", {})
        for key, label in product_map.items():
            if key in lower:
                return label
        return "Unknown"

    def _detect_pdf_product(self, page_text: str) -> str:
        """Detect product type from PDF page header (first 200 chars)."""
        header = page_text[:200].lower()
        if "diesel" in header and "premium" not in header and "magna" not in header:
            return "Diesel"
        elif "87" in header or "magna" in header:
            return "Regular"
        elif "91" in header or "premium" in header:
            return "Premium"
        return ""

    def _parse_price(self, raw) -> float:
        """Parse a price value from any string or numeric format."""
        if raw is None:
            return None
        try:
            # Remove currency symbols, spaces, commas used as thousands sep
            cleaned = str(raw).replace("$", "").replace(" ", "").replace(",", "")
            val = float(cleaned)
            return val if val > 0 else None
        except (ValueError, TypeError):
            return None

    def _convert_price(self, price: float) -> float:
        """Convert price to MXN/L if needed (e.g. Pemex MXN/m³ → MXN/L)."""
        factor = self.config.get("conversion_factor", 1.0)
        return round(price / factor, 6)

    def _extract_numbers(self, line: str) -> list:
        """Extract all decimal numbers from a text line."""
        return [float(x) for x in re.findall(r"\d+\.\d+", line)]

    def _extract_date_filename(self, filename: str) -> str:
        """Extract date from filename using config patterns."""
        cfg = self.config
        patterns = cfg.get("date_patterns", [r"(\d{8})"])
        fmt      = cfg.get("date_format", "%Y%m%d")

        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                try:
                    return datetime.strptime(match.group(1), fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
        return ""

    def _extract_pdf_dates(self, text: str, filename: str) -> tuple:
        """Extract start and end dates from Pemex PDF content."""
        cfg      = self.config
        months   = cfg.get("months_es", {})
        patterns = cfg.get("date_patterns", [])
        sample   = text[:300].lower()

        for pattern in patterns:
            match = re.search(pattern, sample)
            if match:
                try:
                    g = match.groups()
                    if len(g) == 4:
                        # Same month pattern
                        d_start = datetime(int(g[3]), months.get(g[2], 1), int(g[0]))
                        d_end   = datetime(int(g[3]), months.get(g[2], 1), int(g[1]))
                        return d_start.strftime("%Y-%m-%d"), d_end.strftime("%Y-%m-%d")
                    elif len(g) == 6:
                        # Cross-month pattern
                        d_start = datetime(int(g[2]), months.get(g[1], 1), int(g[0]))
                        d_end   = datetime(int(g[5]), months.get(g[4], 1), int(g[3]))
                        return d_start.strftime("%Y-%m-%d"), d_end.strftime("%Y-%m-%d")
                except (ValueError, KeyError):
                    continue

        # Fallback: filename
        fn_pattern = cfg.get("date_pattern_filename", r"(\d{8})_(\d{8})")
        fn_fmt     = cfg.get("date_format_fallback", "%Y%m%d")
        match = re.search(fn_pattern, filename)
        if match:
            try:
                d1 = datetime.strptime(match.group(1), fn_fmt).strftime("%Y-%m-%d")
                d2 = datetime.strptime(match.group(2), fn_fmt).strftime("%Y-%m-%d")
                return d1, d2
            except ValueError:
                pass

        return "", ""

    # ── Output formatting ───────────────────────────────────────────────

    def _to_dataframe(self, records: list, source_file: str) -> pd.DataFrame:
        """
        Convert list of record dicts to a standard schema DataFrame
        with price_flag column added.
        """
        if not records:
            return pd.DataFrame()

        df = pd.DataFrame(records)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["price_mxn_per_l"] = df["price_mxn_per_l"].astype(float)

        # Add price_flag using guardrails from config
        guardrails = self.config.get("price_guardrails", {"min": 15.0, "max": 35.0})
        df["price_flag"] = (
            (df["price_mxn_per_l"] < guardrails["min"]) |
            (df["price_mxn_per_l"] > guardrails["max"])
        )

        flagged = int(df["price_flag"].sum())
        if flagged:
            self.logger.warning(
                "%d price(s) outside guardrail range [%.1f, %.1f] in %s",
                flagged, guardrails["min"], guardrails["max"], source_file
            )

        # Keep only standard output columns that exist
        available = [c for c in OUTPUT_COLUMNS if c in df.columns]
        return df[available].reset_index(drop=True)
