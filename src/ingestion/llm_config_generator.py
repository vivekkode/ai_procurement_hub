"""
llm_config_generator.py
-----------------------
Step 3 of the generic ingestion layer.

When a new supplier file arrives with no existing config, this module:
    1. Reads the file content (any format)
    2. Sends it to Claude API with a structured extraction prompt
    3. Claude identifies: which column is price, terminal, product,
       what unit, what date format, what language
    4. Returns a config dict the GenericParser can use immediately
    5. Saves the config to disk so next time no LLM call is needed

This is what makes the system agentic — a new supplier's file is
processed automatically without any human writing config code.

Roberto's requirement (Mar 13 email):
    "Can your agent read an unstructured email or PDF from Valero
    detailing a sudden weather surcharge, automatically extract that
    variable, update the landed cost, and autonomously route the
    purchase to Pemex instead?"

This module handles the structured part of that — files with columns
and rows. The unstructured part (emails, surcharge notices) is handled
by parse_llm.py (Step 4).

Usage:
    from src.ingestion.llm_config_generator import LLMConfigGenerator

    generator = LLMConfigGenerator()

    # Auto-detect structure and generate config
    config = generator.generate_config(
        filepath="/tmp/g500_20260315.xlsx",
        supplier_name="G500"
    )

    # Config is ready to use immediately
    from src.ingestion.generic_parser import GenericParser
    parser = GenericParser(config)
    df = parser.parse_file("/tmp/g500_20260315.xlsx")

    # Config is also saved to disk automatically
    # Next time G500 sends a file, no LLM call needed
"""

import json
import logging
import os
import re
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

logger = logging.getLogger(__name__)

# Where generated configs are saved
CONFIG_DIR = Path("config/suppliers")

# Standard output fields the LLM must map to
STANDARD_FIELDS = {
    "terminal": "The column containing the terminal, station, or location name",
    "product":  "The column containing the fuel/product type (Regular, Premium, Diesel)",
    "price":    "The column containing the FINAL price after all discounts and taxes",
    "discount": "The column containing any discount applied (optional)",
    "ref_price": "The column containing the base/reference price before discount (optional)",
}


# ---------------------------------------------------------------------------
# LLMConfigGenerator
# ---------------------------------------------------------------------------

class LLMConfigGenerator:
    """
    Generates supplier configs automatically using Claude API.

    One instance can handle any number of unknown suppliers.
    Generated configs are saved to config/suppliers/ automatically.
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def generate_config(self, filepath: str, supplier_name: str) -> dict:
        """
        Generate a supplier config by having Claude analyze the file.

        Args:
            filepath:      Path to any supplier pricing file
            supplier_name: Name to assign to this supplier (e.g. "G500")

        Returns:
            Config dict ready for GenericParser — same structure as
            the manually written configs in config/suppliers/

        Raises:
            ValueError: if file cannot be read or LLM cannot identify structure
        """
        filepath = Path(filepath)
        self.logger.info(
            "Generating config for new supplier '%s' from: %s",
            supplier_name, filepath.name
        )

        # Step 1: Extract a sample of the file content for the LLM
        sample = self._extract_sample(filepath)
        if not sample:
            raise ValueError(f"Could not read file: {filepath}")

        # Step 2: Ask Claude to identify the structure
        self.logger.info("Sending sample to Claude API for structure analysis...")
        llm_response = self._call_claude(sample, supplier_name, filepath.name)

        # Step 3: Parse Claude's response into a config dict
        new_config = self._build_config(llm_response, supplier_name, filepath)

        # Step 4: If an old config exists, MERGE column_maps instead of overwriting
        # This preserves all previously known keywords in case they return
        old_config = self.load_saved_config(supplier_name)
        if old_config:
            old_map = old_config.get("column_map", {})
            new_map = new_config.get("column_map", {})
            merged_map = self._merge_column_maps(old_map, new_map)
            new_config["column_map"] = merged_map

            # Carry forward any uncovered headers log from old config
            old_uncovered = old_config.get("_uncovered_headers", [])
            if old_uncovered:
                new_config.setdefault("_uncovered_headers", [])
                for h in old_uncovered:
                    if h not in new_config["_uncovered_headers"]:
                        new_config["_uncovered_headers"].append(h)

            self.logger.info(
                "Merged column_map with previous config for '%s' — "
                "keywords preserved across format changes",
                supplier_name
            )

        # Step 5: Save merged config to disk for reuse
        self._save_config(new_config, supplier_name)

        self.logger.info(
            "Config generated and saved for '%s' — "
            "future files from this supplier will not need LLM",
            supplier_name
        )
        return new_config

    def config_exists(self, supplier_name: str) -> bool:
        """Check if a saved config already exists for this supplier."""
        config_path = CONFIG_DIR / f"{supplier_name.lower()}_config.py"
        return config_path.exists()

    def load_saved_config(self, supplier_name: str) -> dict:
        """
        Load a previously generated and saved config.
        Returns None if no saved config exists.
        """
        config_path = CONFIG_DIR / f"{supplier_name.lower()}_config.py"
        if not config_path.exists():
            return None

        import importlib.util
        spec = importlib.util.spec_from_file_location("config", config_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return getattr(module, "CONFIG", None)

    def config_still_valid(self, filepath: Path, config: dict,
                           supplier_name: str = None) -> bool:
        """
        Check if a saved config is still compatible with a new file.

        Two checks:
            1. Format — does the file format match what the config expects?
            2. Columns — do at least 50% of the config's column keywords
                         still appear in the new file's headers?

        Returns True if config is still good, False if needs regenerating.
        """
        from src.ingestion.format_detector import detect_format

        # Check 1 — format
        detected_fmt = detect_format(str(filepath)).value
        config_fmt   = config.get("file_format", "")
        if detected_fmt != config_fmt:
            self.logger.warning(
                "Format changed for '%s': config says '%s' but "
                "file is '%s' — regenerating config",
                filepath.name, config_fmt, detected_fmt
            )
            return False

        # Check 2 — column keywords (only for tabular formats)
        if detected_fmt not in ("xlsx", "csv", "xls"):
            return True

        try:
            headers = self._extract_headers(filepath)
        except Exception:
            return True   # cannot read headers, assume still valid

        if not headers:
            return True

        column_map   = config.get("column_map", {})
        total_fields = len(column_map)
        if total_fields == 0:
            return True

        matched = 0
        for field, keywords in column_map.items():
            for keyword in keywords:
                if any(keyword.lower() in h.lower() for h in headers):
                    matched += 1
                    break

        match_rate = matched / total_fields
        if match_rate < 1.0:
            self.logger.warning(
                "%.0f%% of column keywords matched in '%s' — "
                "%d field(s) unmatched, regenerating config to avoid "
                "silent wrong-column extraction",
                match_rate * 100, filepath.name,
                total_fields - matched
            )
            return False

        self.logger.info(
            "Saved config valid for '%s' — 100%% column match",
            filepath.name
        )

        # Even though config is valid, check if new columns were added
        # and enrich the saved config with any new keywords found
        sname = supplier_name or config.get("supplier_name", "unknown")
        self._enrich_config_if_needed(filepath, config, sname)

        return True

    def _enrich_config_if_needed(self, filepath: Path, config: dict,
                                  supplier_name: str):
        """
        Scenario 1 fix — new columns added to file.

        When a supplier adds extra columns to their file (e.g. 6 → 12),
        the 100% match still passes because all original keywords are
        present. But the 6 new columns are invisible to the system.

        This method checks if any new headers appeared that could be
        useful synonyms for existing fields, and silently adds them
        to the saved config's column_map so future files benefit.

        No LLM call needed — just direct header comparison.
        """
        try:
            current_headers = self._extract_headers(filepath)
        except Exception:
            return

        if not current_headers:
            return

        column_map = config.get("column_map", {})

        # Find headers that no keyword currently covers
        uncovered = []
        for header in current_headers:
            covered = False
            for field, keywords in column_map.items():
                if any(kw.lower() in header.lower() for kw in keywords):
                    covered = True
                    break
            if not covered:
                uncovered.append(header)

        if not uncovered:
            return  # all headers already covered

        self.logger.info(
            "%d new column(s) found in '%s' not covered by saved config: %s — "
            "these may be new supplier fields. Config enrichment logged.",
            len(uncovered), filepath.name, uncovered
        )

        # Store uncovered headers in config metadata for human review
        # We do NOT auto-map them — mapping requires knowing which
        # standard field they belong to, which needs human judgement
        # or an LLM call. We flag them instead.
        config.setdefault("_uncovered_headers", [])
        for h in uncovered:
            if h not in config["_uncovered_headers"]:
                config["_uncovered_headers"].append(h)

        # Save updated config with the new metadata
        self._save_config(config, supplier_name)
        self.logger.info(
            "Config updated with uncovered header list — "
            "review '_uncovered_headers' in %s_config.py to decide "
            "if any should be added to column_map",
            supplier_name.lower()
        )

    def _merge_column_maps(self, old_map: dict, new_map: dict) -> dict:
        """
        Scenario 2 fix — merge old and new column_map keyword lists.

        When a supplier changes their columns and the LLM generates a
        new config, we merge instead of overwrite. This preserves all
        keywords from the old config (in case they come back in a future
        file version) while adding any new keywords the LLM found.

        Old keywords come first — they have proven track record.
        New keywords are appended — they handle the current file.
        Duplicates are removed. Order is preserved.

        Example:
            old: {"price": ["costo_final_por_litro", "precio_final"]}
            new: {"price": ["precio_neto_final", "net_price"]}
            merged: {"price": ["costo_final_por_litro", "precio_final",
                                "precio_neto_final", "net_price"]}
        """
        all_fields = set(old_map.keys()) | set(new_map.keys())
        merged = {}

        for field in all_fields:
            old_keywords = old_map.get(field, [])
            new_keywords = new_map.get(field, [])

            # Union — old first, then new, no duplicates
            seen     = set()
            combined = []
            for kw in old_keywords + new_keywords:
                kw_lower = kw.lower()
                if kw_lower not in seen:
                    seen.add(kw_lower)
                    combined.append(kw)

            merged[field] = combined

        return merged

    def _extract_headers(self, filepath: Path) -> list:
        """Extract column header names from a tabular file."""
        suffix = filepath.suffix.lower()
        if suffix in (".xlsx", ".xlsm", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            # Find the header row — the row with most non-None string values
            best_row  = []
            best_score = 0
            for row in rows[:5]:
                strings = [
                    str(c).strip() for c in row
                    if c is not None and str(c).strip()
                ]
                if len(strings) > best_score:
                    best_score = len(strings)
                    best_row   = strings
            return best_row
        elif suffix == ".csv":
            import pandas as pd
            df = pd.read_csv(filepath, nrows=0, encoding="utf-8-sig")
            return list(df.columns)
        return []

    # ── File sampling ───────────────────────────────────────────────────

    def _extract_sample(self, filepath: Path) -> str:
        """
        Extract a representative sample of the file for the LLM.

        We send a small sample — not the whole file — to keep the
        API call fast and cheap. For structured files, the first
        10 rows are enough to identify column names and data patterns.
        """
        suffix = filepath.suffix.lower()

        try:
            if suffix in (".xlsx", ".xlsm", ".xls"):
                return self._sample_excel(filepath)
            elif suffix == ".csv":
                return self._sample_csv(filepath)
            elif suffix == ".html":
                return self._sample_html(filepath)
            elif suffix == ".txt":
                return self._sample_txt(filepath)
            elif suffix == ".pdf":
                return self._sample_pdf(filepath)
            else:
                # Try reading as text
                with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                    return f.read(3000)
        except Exception as e:
            self.logger.error("Could not sample file %s: %s", filepath.name, e)
            return ""

    def _sample_excel(self, filepath: Path) -> str:
        """Extract first 10 rows from Excel as readable text."""
        import openpyxl
        wb = openpyxl.load_workbook(str(filepath), data_only=True)

        # Try to get the first non-empty sheet
        ws = wb.active
        lines = [f"SHEET NAME: {ws.title}"]
        lines.append("ROWS (first 10):")

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            row_str = " | ".join(str(c) if c is not None else "" for c in row)
            lines.append(f"  Row {i}: {row_str}")

        return "\n".join(lines)

    def _sample_csv(self, filepath: Path) -> str:
        """Extract first 10 rows from CSV as readable text."""
        import pandas as pd
        df = pd.read_csv(filepath, nrows=10, encoding="utf-8-sig")
        lines = ["CSV COLUMNS: " + " | ".join(df.columns)]
        lines.append("FIRST 5 ROWS:")
        for _, row in df.head(5).iterrows():
            lines.append("  " + " | ".join(str(v) for v in row))
        return "\n".join(lines)

    def _sample_html(self, filepath: Path) -> str:
        """Extract table structure from HTML."""
        from bs4 import BeautifulSoup
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            soup = BeautifulSoup(f, "html.parser")

        lines = ["HTML STRUCTURE:"]
        # Get all tables
        for i, table in enumerate(soup.find_all("table")[:2]):
            lines.append(f"\nTable {i+1}:")
            for j, row in enumerate(table.find_all("tr")[:5]):
                cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                lines.append(f"  Row {j}: {' | '.join(cells)}")

        # Get card-style divs if no tables
        if not soup.find("table"):
            divs = soup.find_all("div", class_=True)[:5]
            for d in divs:
                lines.append(f"  Div class={d.get('class')}: {d.get_text(strip=True)[:100]}")

        return "\n".join(lines)

    def _sample_txt(self, filepath: Path) -> str:
        """Return first 50 lines of text file."""
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()[:50]
        return "TXT FILE CONTENT (first 50 lines):\n" + "".join(lines)

    def _sample_pdf(self, filepath: Path) -> str:
        """Extract first 500 chars from PDF."""
        import pdfminer.high_level
        text = pdfminer.high_level.extract_text(str(filepath))
        return "PDF CONTENT (first 500 chars):\n" + (text or "")[:500]

    # ── Claude API call ──────────────────────────────────────────────────

    def _call_claude(self, sample: str, supplier_name: str,
                     filename: str) -> dict:
        """
        Call Claude API to analyze file structure and identify fields.

        Returns a dict with Claude's analysis of the file structure.
        """
        import urllib.request

        prompt = self._build_prompt(sample, supplier_name, filename)

        payload = json.dumps({
            "model": "claude-sonnet-4-5",
            "max_tokens": 1000,
            "messages": [{"role": "user", "content": prompt}]
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": os.getenv("ANTHROPIC_API_KEY", ""),
                "anthropic-version": "2023-06-01",
            },
            method="POST"
        )

        try:
            with urllib.request.urlopen(req) as response:
                data = json.loads(response.read())
                raw_text = data["content"][0]["text"]
                return self._parse_llm_response(raw_text)
        except Exception as e:
            self.logger.error("Claude API call failed: %s", e)
            raise ValueError(f"LLM API call failed: {e}")

    def _build_prompt(self, sample: str, supplier_name: str,
                      filename: str) -> str:
        """Build the structured extraction prompt for Claude."""
        fields_desc = "\n".join(
            f'  - "{field}": {desc}'
            for field, desc in STANDARD_FIELDS.items()
        )

        return f"""You are a data extraction expert analyzing a fuel supplier pricing file.

Supplier name: {supplier_name}
Filename: {filename}

FILE CONTENT SAMPLE:
{sample}

Analyze this file and return a JSON object with EXACTLY these fields.
Return ONLY the JSON — no explanation, no markdown, no backticks.

{{
  "file_format": "xlsx|csv|html|txt|pdf",
  "sheet_name": "exact sheet name if Excel, null otherwise",
  "header_row": 0,
  "data_start_row": 1,
  "date_source": "filename|title_row|content",
  "date_pattern": "regex to extract date from filename or title row",
  "date_format": "strptime format like %d/%m/%Y or %Y%m%d",
  "structure": "xlsx_table|html_cards|txt_labelled_rows|pdf_raw_text|csv_table",
  "price_unit": "MXN_per_liter|MXN_per_m3|USD_per_gallon",
  "conversion_factor": 1.0,
  "column_map": {{
    "terminal":  ["list", "of", "possible", "column", "name", "keywords"],
    "product":   ["list", "of", "possible", "column", "name", "keywords"],
    "price":     ["list", "of", "possible", "column", "name", "keywords"],
    "discount":  ["list", "of", "possible", "column", "name", "keywords"],
    "ref_price": ["list", "of", "possible", "column", "name", "keywords"]
  }},
  "product_map": {{
    "keyword_in_product_column": "Regular|Premium|Diesel"
  }},
  "terminal_pattern": "regex to parse terminal string into id and name",
  "contract_type": "inferred contract type or empty string",
  "price_guardrails": {{"min": 15.0, "max": 35.0}},
  "confidence": "high|medium|low",
  "notes": "any important observations about this file structure"
}}

Rules:
- column_map keywords must be lowercase substrings of the actual column names
- product_map keys must be lowercase substrings found in the product column values
- price_guardrails should be 15.0-35.0 for MXN/L (standard Mexican fuel range)
- If price is in MXN/m³ set conversion_factor to 1000.0
- If price is in USD/gallon set conversion_factor to 0.2642 * exchange_rate (use 0.0531 as default)
- Set confidence based on how clearly you can identify the price column"""

    def _parse_llm_response(self, raw_text: str) -> dict:
        """Parse Claude's JSON response, handling any extra text."""
        # Strip markdown code blocks if present
        text = re.sub(r"```json\s*|\s*```", "", raw_text).strip()

        # Find JSON object in response
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if not match:
            raise ValueError(f"No JSON found in LLM response: {raw_text[:200]}")

        try:
            return json.loads(match.group())
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON from LLM: {e}")

    # ── Config building ──────────────────────────────────────────────────

    def _build_config(self, llm_response: dict, supplier_name: str,
                      filepath: Path) -> dict:
        """
        Convert LLM's JSON response into a full config dict
        compatible with GenericParser.
        """
        confidence = llm_response.get("confidence", "low")
        if confidence == "low":
            self.logger.warning(
                "LLM has LOW confidence in config for '%s' — "
                "recommend human review before using in production",
                supplier_name
            )

        config = {
            "supplier_name":   supplier_name,
            "file_format":     llm_response.get("file_format", "xlsx"),
            "file_extensions": [f".{llm_response.get('file_format', 'xlsx')}"],
            "contract_type":   llm_response.get("contract_type", ""),
            "country":         "MX",

            # Date extraction
            "date_source":   llm_response.get("date_source", "filename"),
            "date_patterns": [llm_response.get("date_pattern", r"(\d{8})")],
            "date_format":   llm_response.get("date_format", "%Y%m%d"),

            # Document structure
            "structure":   llm_response.get("structure", "xlsx_table"),
            "sheet_name":  llm_response.get("sheet_name"),
            "sheet_index": 0,
            "header_row":  llm_response.get("header_row", 0),
            "data_start":  llm_response.get("data_start_row", 1),

            # Column mapping — this is the key output of the LLM
            "column_map": llm_response.get("column_map", {}),

            # Price
            "price_unit":        llm_response.get("price_unit", "MXN_per_liter"),
            "conversion_factor": float(llm_response.get("conversion_factor", 1.0)),
            "price_guardrails":  llm_response.get(
                "price_guardrails", {"min": 15.0, "max": 35.0}
            ),

            # Product normalization
            "product_map": llm_response.get("product_map", {
                "regular": "Regular", "87": "Regular", "magna": "Regular",
                "premium": "Premium", "91": "Premium",
                "diesel":  "Diesel",
            }),

            # Terminal parsing
            "terminal_pattern": llm_response.get("terminal_pattern", r"^(.+?)\s*-\s*(.+)$"),
            "terminal_groups": {"terminal_id": 1, "terminal_name": 2},

            # Metadata
            "_generated_by": "llm_config_generator",
            "_confidence":   confidence,
            "_notes":        llm_response.get("notes", ""),
        }

        return config

    # ── Config persistence ───────────────────────────────────────────────

    def _save_config(self, config: dict, supplier_name: str):
        """
        Save generated config to disk as a Python file.

        Saved configs are identical in structure to the manually written
        ones (valero_config.py, exxon_config.py, etc.) so they work
        with GenericParser without any special handling.
        """
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        config_path = CONFIG_DIR / f"{supplier_name.lower()}_config.py"

        # Format config as a readable Python file
        lines = [
            f'"""',
            f'{supplier_name.lower()}_config.py',
            f'---------------------------',
            f'Auto-generated config for {supplier_name}.',
            f'Generated by: llm_config_generator.py',
            f'Confidence: {config.get("_confidence", "unknown")}',
            f'Notes: {config.get("_notes", "")}',
            f'',
            f'Review the column_map before using in production.',
            f'"""',
            f'',
            f'CONFIG = {json.dumps(config, indent=4, ensure_ascii=False)}',
        ]

        with open(config_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        self.logger.info("Config saved to: %s", config_path)


# ---------------------------------------------------------------------------
# Convenience function — main entry point
# ---------------------------------------------------------------------------

def auto_onboard_supplier(filepath: str, supplier_name: str,
                           force_regenerate: bool = False) -> dict:
    """
    Full auto-onboarding flow for a new supplier.

    Checks if a config already exists. If yes, loads and returns it.
    If no, generates one using Claude and saves it for future use.

    Args:
        filepath:         Path to any supplier pricing file
        supplier_name:    Name for this supplier (e.g. "G500")
        force_regenerate: If True, regenerate config even if one exists

    Returns:
        Config dict ready for GenericParser

    Example:
        config = auto_onboard_supplier(
            filepath="data/raw/g500/g500_20260315.xlsx",
            supplier_name="G500"
        )
        from src.ingestion.generic_parser import GenericParser
        parser = GenericParser(config)
        df = parser.parse_file("data/raw/g500/g500_20260315.xlsx")
    """
    generator = LLMConfigGenerator()
    filepath  = Path(filepath)

    if not force_regenerate and generator.config_exists(supplier_name):
        saved_config = generator.load_saved_config(supplier_name)

        if saved_config and generator.config_still_valid(filepath, saved_config, supplier_name):
            # Config exists and still matches the file — use it directly
            logger.info(
                "Using saved config for '%s' — no LLM call needed",
                supplier_name
            )
            return saved_config

        # Config exists but is stale (format or columns changed)
        logger.warning(
            "Saved config for '%s' is outdated — regenerating via LLM",
            supplier_name
        )

    # Generate fresh config via LLM
    return generator.generate_config(str(filepath), supplier_name)


# ---------------------------------------------------------------------------
# CLI — test on any file
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s"
    )

    if len(sys.argv) < 3:
        print("Usage: python llm_config_generator.py <filepath> <supplier_name>")
        print()
        print("Example:")
        print("  python llm_config_generator.py data/raw/g500/g500_20260315.xlsx G500")
        sys.exit(1)

    filepath      = sys.argv[1]
    supplier_name = sys.argv[2]

    print(f"\nAuto-onboarding supplier: {supplier_name}")
    print(f"File: {filepath}")
    print("-" * 50)

    try:
        config = auto_onboard_supplier(filepath, supplier_name)
        print("\nGenerated config:")
        print(json.dumps(config, indent=2, ensure_ascii=False))

        # Immediately test with generic parser
        print("\nTesting generic parser with generated config...")
        sys.path.insert(0, ".")
        from src.ingestion.generic_parser import GenericParser
        parser = GenericParser(config)
        df = parser.parse_file(filepath)

        if df.empty:
            print("WARNING: Parser returned empty DataFrame")
            print("Config may need manual review")
        else:
            print(f"SUCCESS: {len(df)} records extracted")
            print(f"Columns: {list(df.columns)}")
            print(f"Sample:\n{df.head(3).to_string()}")

    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)