"""
tests/test_parsers.py
---------------------
Tests for the generic parser and all 4 supplier configs.

Replaces the original test suite which tested the individual parsers
(parse_valero, parse_exxon, parse_marathon, parse_pemex).

Run with:
    python tests/test_parsers.py
"""

import os
import sys
import shutil
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.ingestion.generic_parser       import GenericParser
from src.ingestion.format_detector      import detect_format, FileFormat
from src.ingestion.llm_config_generator import LLMConfigGenerator
from src.ingestion.parse_llm            import UnstructuredParser, SurchargeEvent

from config.suppliers.valero_config   import CONFIG as VALERO_CONFIG
from config.suppliers.exxon_config    import CONFIG as EXXON_CONFIG
from config.suppliers.marathon_config import CONFIG as MARATHON_CONFIG
from config.suppliers.pemex_config    import CONFIG as PEMEX_CONFIG

ROOT          = Path(__file__).resolve().parents[1]
VALERO_FILE   = ROOT / "data/raw/valero/valero_20240101.html"
EXXON_FILE    = ROOT / "data/raw/exxon/exxon_20240101.xlsx"
MARATHON_FILE = ROOT / "data/raw/marathon/marathon_20240101_MX-IT-Chihuahua.txt"
PEMEX_FILE    = ROOT / "data/raw/pemex/reportes_pdf/pemex_tar_20240101_20240105.pdf"

def tmp():
    return tempfile.mkdtemp()


# ── FORMAT DETECTOR ───────────────────────────────────────────────────────────

class TestFormatDetector(unittest.TestCase):

    def test_html_detected(self):
        self.assertEqual(detect_format(str(VALERO_FILE)), FileFormat.HTML)
        print(f"  ✓ HTML detected")

    def test_xlsx_detected(self):
        self.assertEqual(detect_format(str(EXXON_FILE)), FileFormat.XLSX)
        print(f"  ✓ XLSX detected")

    def test_txt_detected(self):
        self.assertEqual(detect_format(str(MARATHON_FILE)), FileFormat.TXT)
        print(f"  ✓ TXT detected")

    def test_pdf_detected(self):
        self.assertEqual(detect_format(str(PEMEX_FILE)), FileFormat.PDF)
        print(f"  ✓ PDF detected")

    def test_pdf_renamed_to_txt_detected_as_pdf(self):
        tmpdir = tmp()
        try:
            fake = Path(tmpdir) / "tricky.txt"
            fake.write_bytes(PEMEX_FILE.read_bytes())
            self.assertEqual(detect_format(str(fake)), FileFormat.PDF)
            print(f"  ✓ Magic bytes win: PDF named .txt → pdf")
        finally:
            shutil.rmtree(tmpdir)

    def test_empty_file_returns_unknown(self):
        tmpdir = tmp()
        try:
            empty = Path(tmpdir) / "empty.xlsx"
            empty.write_bytes(b"")
            self.assertEqual(detect_format(str(empty)), FileFormat.UNKNOWN)
            print(f"  ✓ Empty file → unknown")
        finally:
            shutil.rmtree(tmpdir)

    def test_nonexistent_file_returns_unknown(self):
        self.assertEqual(detect_format("/tmp/no_such_file_xyz.pdf"), FileFormat.UNKNOWN)
        print(f"  ✓ Nonexistent file → unknown")


# ── VALERO ────────────────────────────────────────────────────────────────────

class TestValeroGenericParser(unittest.TestCase):

    def setUp(self):
        self.parser = GenericParser(VALERO_CONFIG)

    def test_normal_file_row_count(self):
        df = self.parser.parse_file(str(VALERO_FILE))
        self.assertEqual(len(df), 29)
        self.assertEqual(df["terminal_id"].nunique(), 9)
        print(f"  ✓ Normal: {len(df)} rows, {df['terminal_id'].nunique()} terminals")

    def test_output_schema_has_all_columns(self):
        df = self.parser.parse_file(str(VALERO_FILE))
        for col in ["date","supplier","terminal_id","terminal_name","state",
                    "country","product_type","price_mxn_per_l",
                    "contract_type","source_file","price_flag"]:
            self.assertIn(col, df.columns)
        print(f"  ✓ Schema: all 11 columns present")

    def test_supplier_name(self):
        df = self.parser.parse_file(str(VALERO_FILE))
        self.assertTrue((df["supplier"] == "Valero").all())
        print(f"  ✓ Supplier: Valero")

    def test_products_normalised(self):
        df = self.parser.parse_file(str(VALERO_FILE))
        self.assertTrue(df["product_type"].isin({"Regular","Premium","Diesel","Unknown"}).all())
        print(f"  ✓ Products: {sorted(df['product_type'].unique())}")

    def test_price_guardrail_flags_bad_price(self):
        tmpdir = tmp()
        try:
            mod = VALERO_FILE.read_text(encoding="utf-8", errors="replace").replace("20.987243","0.01")
            f = Path(tmpdir) / "valero_20240101.html"
            f.write_text(mod, encoding="utf-8")
            df = self.parser.parse_file(str(f))
            self.assertGreater(df["price_flag"].sum(), 0)
            print(f"  ✓ Price guardrail: flagged {df['price_flag'].sum()} bad price(s)")
        finally:
            shutil.rmtree(tmpdir)

    def test_extra_html_wrapper_ignored(self):
        tmpdir = tmp()
        try:
            mod = VALERO_FILE.read_text(encoding="utf-8", errors="replace")
            mod = mod.replace("<body>","<body><div id='wrapper'>").replace("</body>","</div></body>")
            f = Path(tmpdir) / "valero_20240101.html"
            f.write_text(mod, encoding="utf-8")
            df = self.parser.parse_file(str(f))
            self.assertEqual(len(df), 29)
            print(f"  ✓ Extra HTML wrapper: still {len(df)} rows")
        finally:
            shutil.rmtree(tmpdir)

    def test_new_terminal_picked_up(self):
        tmpdir = tmp()
        try:
            new_card = """<div class="terminal-card">
              <div class="card-header"><span>Tijuana, BCN, MX - TMX000099</span></div>
              <table><tbody><tr>
                <td>Sin Marca</td><td>Diesel</td>
                <td>01/01/2024 12:00:00 AM</td><td>21.500000</td>
                <td>21.500000</td><td>-------</td>
              </tr></tbody></table></div>"""
            mod = VALERO_FILE.read_text(encoding="utf-8", errors="replace").replace("</body>", new_card+"</body>")
            f = Path(tmpdir) / "valero_20240101.html"
            f.write_text(mod, encoding="utf-8")
            df = self.parser.parse_file(str(f))
            self.assertEqual(df["terminal_id"].nunique(), 10)
            self.assertIn("TMX000099", df["terminal_id"].values)
            print(f"  ✓ New terminal: TMX000099 detected ({df['terminal_id'].nunique()} total)")
        finally:
            shutil.rmtree(tmpdir)

    def test_intraday_prices_preserved(self):
        df = self.parser.parse_file(str(VALERO_FILE))
        t74 = df[(df["terminal_id"]=="T74-TX-2709") & (df["product_type"]=="Regular")]
        self.assertGreater(len(t74), 1)
        print(f"  ✓ Intraday prices: T74-TX-2709 Regular has {len(t74)} snapshots")

    def test_missing_file_returns_empty(self):
        self.assertTrue(self.parser.parse_file("data/raw/valero/no_file.html").empty)
        print(f"  ✓ Missing file → empty")

    def test_full_folder_row_count(self):
        df = self.parser.parse_folder("data/raw/valero")
        self.assertEqual(len(df), 16849)
        self.assertEqual(df["price_flag"].sum(), 0)
        print(f"  ✓ Full folder: {len(df):,} rows, 0 flagged")


# ── EXXON ─────────────────────────────────────────────────────────────────────

class TestExxonGenericParser(unittest.TestCase):

    def setUp(self):
        self.parser = GenericParser(EXXON_CONFIG)

    def test_normal_file_row_count(self):
        df = self.parser.parse_file(str(EXXON_FILE))
        self.assertEqual(len(df), 9)
        self.assertEqual(df["terminal_id"].nunique(), 3)
        print(f"  ✓ Normal: {len(df)} rows, {df['terminal_id'].nunique()} terminals")

    def test_date_from_title_row(self):
        df = self.parser.parse_file(str(EXXON_FILE))
        self.assertEqual(str(df["date"].iloc[0].date()), "2024-01-01")
        print(f"  ✓ Date from title: {df['date'].iloc[0].date()}")

    def test_column_reorder_resilience(self):
        import openpyxl
        tmpdir = tmp()
        try:
            wb = openpyxl.load_workbook(str(EXXON_FILE), data_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Precios Exxon"
            ws2.append([rows[0][0]] + [None]*5)
            ws2.append([None]*6)
            # Swap Terminal and Producto columns
            ws2.append([rows[2][2],rows[2][0],rows[2][1],rows[2][3],rows[2][4],rows[2][5]])
            for row in rows[3:]:
                if any(row):
                    ws2.append([row[2],row[0],row[1],row[3],row[4],row[5]])
            f = Path(tmpdir) / "exxon_20240101.xlsx"
            wb2.save(str(f))
            df = self.parser.parse_file(str(f))
            self.assertEqual(len(df), 9)
            print(f"  ✓ Column reorder: still {len(df)} rows")
        finally:
            shutil.rmtree(tmpdir)

    def test_wrong_sheet_name_fallback(self):
        import openpyxl
        tmpdir = tmp()
        try:
            wb = openpyxl.load_workbook(str(EXXON_FILE), data_only=True)
            wb.active.title = "Nombre_Nuevo"
            f = Path(tmpdir) / "exxon_20240101.xlsx"
            wb.save(str(f))
            df = self.parser.parse_file(str(f))
            self.assertFalse(df.empty)
            print(f"  ✓ Sheet rename: fallback to index 0, got {len(df)} rows")
        finally:
            shutil.rmtree(tmpdir)

    def test_invoice_price_used(self):
        df = self.parser.parse_file(str(EXXON_FILE))
        self.assertTrue((df["price_mxn_per_l"] < 22).all())
        print(f"  ✓ Invoice price: max={df['price_mxn_per_l'].max():.2f} MXN/L")

    def test_missing_file_returns_empty(self):
        self.assertTrue(self.parser.parse_file("data/raw/exxon/no_file.xlsx").empty)
        print(f"  ✓ Missing file → empty")

    def test_full_folder_row_count(self):
        df = self.parser.parse_folder("data/raw/exxon")
        self.assertEqual(len(df), 4905)
        self.assertEqual(df["price_flag"].sum(), 0)
        print(f"  ✓ Full folder: {len(df):,} rows, 0 flagged")


# ── MARATHON ──────────────────────────────────────────────────────────────────

class TestMarathonGenericParser(unittest.TestCase):

    def setUp(self):
        self.parser = GenericParser(MARATHON_CONFIG)

    def test_normal_file_row_count(self):
        df = self.parser.parse_file(str(MARATHON_FILE))
        self.assertEqual(len(df), 3)
        self.assertEqual(set(df["product_type"]), {"Regular","Premium","Diesel"})
        print(f"  ✓ Normal: {len(df)} rows — Regular, Premium, Diesel")

    def test_terminal_from_filename(self):
        df = self.parser.parse_file(str(MARATHON_FILE))
        self.assertEqual(df["terminal_id"].iloc[0], "MX-IT-Chihuahua")
        self.assertEqual(df["terminal_name"].iloc[0], "Chihuahua")
        print(f"  ✓ Terminal: {df['terminal_id'].iloc[0]}")

    def test_date_from_filename(self):
        df = self.parser.parse_file(str(MARATHON_FILE))
        self.assertEqual(str(df["date"].iloc[0].date()), "2024-01-01")
        print(f"  ✓ Date: {df['date'].iloc[0].date()}")

    def test_invoice_price_includes_iva(self):
        df = self.parser.parse_file(str(MARATHON_FILE))
        self.assertTrue((df["price_mxn_per_l"] > 18).all())
        print(f"  ✓ Invoice price (with IVA): {df['price_mxn_per_l'].tolist()}")

    def test_us_terminal_country(self):
        tmpdir = tmp()
        try:
            content = MARATHON_FILE.read_text(encoding="utf-8", errors="replace")
            f = Path(tmpdir) / "marathon_20240101_US-IT-El_Paso.txt"
            f.write_text(content, encoding="utf-8")
            df = self.parser.parse_file(str(f))
            self.assertEqual(df["country"].iloc[0], "US")
            self.assertEqual(df["terminal_name"].iloc[0], "El Paso")
            print(f"  ✓ US terminal: country=US, name=El Paso")
        finally:
            shutil.rmtree(tmpdir)

    def test_missing_invoice_price_returns_empty(self):
        tmpdir = tmp()
        try:
            content = MARATHON_FILE.read_text(encoding="utf-8", errors="replace")
            mod = "\n".join(l for l in content.splitlines() if not l.strip().startswith("Invoice Price"))
            f = Path(tmpdir) / "marathon_20240101_MX-IT-Chihuahua.txt"
            f.write_text(mod, encoding="utf-8")
            self.assertTrue(self.parser.parse_file(str(f)).empty)
            print(f"  ✓ Missing invoice price → empty")
        finally:
            shutil.rmtree(tmpdir)

    def test_full_folder_row_count(self):
        df = self.parser.parse_folder("data/raw/marathon")
        self.assertEqual(len(df), 16350)
        self.assertEqual(df["price_flag"].sum(), 0)
        print(f"  ✓ Full folder: {len(df):,} rows, 0 flagged")


# ── PEMEX ─────────────────────────────────────────────────────────────────────

class TestPemexGenericParser(unittest.TestCase):

    def setUp(self):
        self.parser = GenericParser(PEMEX_CONFIG)

    def test_normal_file_row_count(self):
        df = self.parser.parse_file(str(PEMEX_FILE))
        self.assertEqual(len(df), 234)
        self.assertEqual(df["terminal_id"].nunique(), 78)
        print(f"  ✓ Normal: {len(df)} rows, {df['terminal_id'].nunique()} regions")

    def test_unit_conversion_m3_to_liter(self):
        df = self.parser.parse_file(str(PEMEX_FILE))
        self.assertTrue((df["price_mxn_per_l"] > 15).all())
        self.assertTrue((df["price_mxn_per_l"] < 35).all())
        self.assertEqual(df["price_flag"].sum(), 0)
        sample = df[df["terminal_id"]=="CHIHUAHUA"].iloc[0]
        print(f"  ✓ Unit conversion: Chihuahua = {sample['price_mxn_per_l']:.4f} MXN/L")

    def test_all_product_types(self):
        df = self.parser.parse_file(str(PEMEX_FILE))
        counts = df["product_type"].value_counts()
        self.assertEqual(counts.get("Regular",0), 78)
        self.assertEqual(counts.get("Premium",0), 78)
        self.assertEqual(counts.get("Diesel",0),  78)
        print(f"  ✓ Products: Regular=78, Premium=78, Diesel=78")

    def test_spanish_date_parsed(self):
        df = self.parser.parse_file(str(PEMEX_FILE))
        self.assertEqual(str(df["date"].iloc[0].date()), "2024-01-01")
        print(f"  ✓ Spanish date → {df['date'].iloc[0].date()}")

    def test_missing_file_returns_empty(self):
        self.assertTrue(self.parser.parse_file("data/raw/pemex/reportes_pdf/no_file.pdf").empty)
        print(f"  ✓ Missing file → empty")

    def test_full_folder_row_count(self):
        df = self.parser.parse_folder("data/raw/pemex/reportes_pdf")
        self.assertEqual(len(df), 25974)
        self.assertEqual(df["price_flag"].sum(), 0)
        print(f"  ✓ Full folder: {len(df):,} rows, 0 flagged")


# ── LLM CONFIG GENERATOR ──────────────────────────────────────────────────────

class TestLLMConfigGenerator(unittest.TestCase):

    def setUp(self):
        self.gen = LLMConfigGenerator()

    def test_100_percent_match_valid(self):
        valid = self.gen.config_still_valid(Path(str(EXXON_FILE)), EXXON_CONFIG, "Exxon")
        self.assertTrue(valid)
        print(f"  ✓ 100% match: config valid")

    def test_format_change_detected(self):
        import csv
        tmpdir = tmp()
        try:
            f = Path(tmpdir) / "prices.csv"
            with open(f,"w",newline="") as fh:
                csv.writer(fh).writerow(["terminal","producto","precio"])
            valid = self.gen.config_still_valid(f, EXXON_CONFIG, "Exxon")
            self.assertFalse(valid)
            print(f"  ✓ Format change: xlsx config vs csv → stale")
        finally:
            shutil.rmtree(tmpdir)

    def test_single_column_change_triggers_regeneration(self):
        import openpyxl
        tmpdir = tmp()
        try:
            wb = openpyxl.load_workbook(str(EXXON_FILE), data_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Precios Exxon"
            ws2.append([rows[0][0]]+[None]*5)
            ws2.append([None]*6)
            ws2.append(["Terminal","Categoría Cuenta","Producto",
                        "Precio Referencia Industrial (MXN/L)",
                        "Descuento (MXN/L)", "Precio_Definitivo_Nuevo"])
            for row in rows[3:]:
                if any(row): ws2.append(list(row))
            f = Path(tmpdir) / "exxon_20240101.xlsx"
            wb2.save(str(f))
            valid = self.gen.config_still_valid(f, EXXON_CONFIG, "Exxon")
            self.assertFalse(valid)
            print(f"  ✓ Single column change → regenerate")
        finally:
            shutil.rmtree(tmpdir)

    def test_merge_preserves_old_keywords(self):
        old = {"price": ["costo_final","precio_final"], "terminal": ["estacion"]}
        new = {"price": ["precio_neto","net_price"],    "terminal": ["punto_venta"]}
        merged = self.gen._merge_column_maps(old, new)
        self.assertIn("costo_final",  merged["price"])
        self.assertIn("precio_final", merged["price"])
        self.assertIn("precio_neto",  merged["price"])
        self.assertIn("net_price",    merged["price"])
        self.assertEqual(len(merged["price"]), len(set(merged["price"])))
        print(f"  ✓ Merge: {merged['price']}")


# ── UNSTRUCTURED PARSER ───────────────────────────────────────────────────────

class TestUnstructuredParser(unittest.TestCase):

    def setUp(self):
        self.parser = UnstructuredParser()
        self._events = None

    def _get_events(self):
        if self._events is None:
            json_str = '''[
                {"supplier":"Valero","terminal":"Nuevo Laredo, TMS","product":"Regular",
                 "surcharge_per_l":0.85,"effective_from":"2026-03-27",
                 "effective_to":"2026-04-04","reason":"Highway 85 closure","confidence":"high"},
                {"supplier":"Valero","terminal":"Nuevo Laredo, TMS","product":"Premium",
                 "surcharge_per_l":0.85,"effective_from":"2026-03-27",
                 "effective_to":"2026-04-04","reason":"Highway 85 closure","confidence":"high"}
            ]'''
            self._events = self.parser._parse_response(json_str, "test.txt")
            self.parser.save_events(self._events)
        return self._events

    def test_events_extracted(self):
        events = self._get_events()
        self.assertEqual(len(events), 2)
        print(f"  ✓ Extracted {len(events)} events")

    def test_supplier_normalised(self):
        self.assertEqual(self._get_events()[0].supplier, "Valero")
        print(f"  ✓ Supplier: Valero")

    def test_products_normalised(self):
        products = {e.product for e in self._get_events()}
        self.assertEqual(products, {"Regular","Premium"})
        print(f"  ✓ Products: {products}")

    def test_surcharge_applied_correct_order(self):
        self._get_events()
        s = self.parser.get_surcharge_for_order("Valero","Nuevo Laredo, TMS","Regular","2026-03-28")
        self.assertEqual(s, 0.85)
        print(f"  ✓ Surcharge applied: +{s} MXN/L")

    def test_diesel_not_affected(self):
        self._get_events()
        s = self.parser.get_surcharge_for_order("Valero","Nuevo Laredo, TMS","Diesel","2026-03-28")
        self.assertEqual(s, 0.0)
        print(f"  ✓ Diesel unaffected: {s}")

    def test_different_terminal_not_affected(self):
        self._get_events()
        s = self.parser.get_surcharge_for_order("Valero","Altamira, TMS","Regular","2026-03-28")
        self.assertEqual(s, 0.0)
        print(f"  ✓ Altamira unaffected: {s}")

    def test_surcharge_expired(self):
        self._get_events()
        s = self.parser.get_surcharge_for_order("Valero","Nuevo Laredo, TMS","Regular","2026-04-10")
        self.assertEqual(s, 0.0)
        print(f"  ✓ Expired surcharge: {s}")

    def test_recommendation_changes(self):
        self._get_events()
        prices = {
            ("Valero","Nuevo Laredo, TMS","Regular"): 20.14,
            ("Marathon","MX-IT-Chihuahua","Regular"): 20.41,
            ("Pemex","NUEVO LEON","Regular"):         21.32,
        }
        adjusted = {k: v + self.parser.get_surcharge_for_order(*k,"2026-03-28") for k,v in prices.items()}
        cheapest = min(adjusted, key=adjusted.get)
        self.assertEqual(cheapest[0], "Marathon")
        print(f"  ✓ Recommendation: {cheapest[0]} at {adjusted[cheapest]:.2f} MXN/L")

    def test_negative_surcharge_rejected(self):
        bad = '[{"supplier":"Valero","terminal":"MTY","product":"Regular","surcharge_per_l":-0.5,"effective_from":"2026-03-27","effective_to":"2026-04-04","reason":"test","confidence":"high"}]'
        self.assertEqual(len(self.parser._parse_response(bad,"test.txt")), 0)
        print(f"  ✓ Negative surcharge rejected")

    def test_empty_response_returns_empty(self):
        self.assertEqual(self.parser._parse_response("[]","test.txt"), [])
        print(f"  ✓ Empty response → []")


# ── FULL PIPELINE ─────────────────────────────────────────────────────────────

class TestFullPipeline(unittest.TestCase):

    def test_combined_row_count_matches_original(self):
        """Full pipeline must produce exactly 64,078 rows."""
        import pandas as pd
        frames = []
        for name, config, folder in [
            ("valero",   VALERO_CONFIG,   "data/raw/valero"),
            ("exxon",    EXXON_CONFIG,    "data/raw/exxon"),
            ("marathon", MARATHON_CONFIG, "data/raw/marathon"),
            ("pemex",    PEMEX_CONFIG,    "data/raw/pemex/reportes_pdf"),
        ]:
            df = GenericParser(config).parse_folder(folder)
            frames.append(df)
            print(f"    {name.capitalize():<12} {len(df):>7,} rows")

        combined = pd.concat(frames, ignore_index=True)
        self.assertEqual(len(combined), 64078)
        self.assertEqual(combined["price_flag"].sum(), 0)
        print(f"  ✓ Total: {len(combined):,} rows | 0 flagged")


# ── RUNNER ────────────────────────────────────────────────────────────────────

def run_tests():
    test_classes = [
        ("FORMAT DETECTOR",      TestFormatDetector),
        ("VALERO   (HTML)",      TestValeroGenericParser),
        ("EXXON    (Excel)",     TestExxonGenericParser),
        ("MARATHON (TXT)",       TestMarathonGenericParser),
        ("PEMEX    (PDF)",       TestPemexGenericParser),
        ("LLM CONFIG GENERATOR", TestLLMConfigGenerator),
        ("UNSTRUCTURED PARSER",  TestUnstructuredParser),
        ("FULL PIPELINE",        TestFullPipeline),
    ]

    passed = failed = errors = 0

    for label, cls in test_classes:
        print()
        print(f"{'='*60}")
        print(f"  {label}")
        print(f"{'='*60}")
        for test in unittest.TestLoader().loadTestsFromTestCase(cls):
            try:
                test.debug()
                passed += 1
            except unittest.SkipTest as e:
                print(f"  ⊘ {test._testMethodName}: SKIPPED — {e}")
            except AssertionError as e:
                print(f"  ✗ {test._testMethodName}: FAILED — {e}")
                failed += 1
            except Exception as e:
                print(f"  ✗ {test._testMethodName}: ERROR — {e}")
                errors += 1

    print()
    print(f"{'='*60}")
    print(f"  RESULTS")
    print(f"{'='*60}")
    print(f"  Passed : {passed}")
    print(f"  Failed : {failed}")
    print(f"  Errors : {errors}")
    print(f"  Total  : {passed + failed + errors}")
    print(f"{'='*60}")
    print()
    if failed == 0 and errors == 0:
        print("  ALL TESTS PASSED ✓")
    else:
        print("  SOME TESTS FAILED — check output above")
    print()


if __name__ == "__main__":
    run_tests()